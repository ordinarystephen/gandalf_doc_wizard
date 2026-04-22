# doc_qa/qa/batch.py
"""Batch runner: N documents × M questions → answer grid + trace log."""

import logging
from typing import Any, Callable, Dict, List, Optional, Tuple

import pandas as pd

logger = logging.getLogger(__name__)


def _answer_for_doc(question: str, filename: str, doc_id: str, chain) -> "AnswerResult":
    """Answer one question against one document's FAISS index."""
    from doc_qa.retrieval.vectorstore import load_index
    from doc_qa.retrieval.retriever import retrieve
    from doc_qa.qa.chain import answer_question

    index, metadata_dict, position_map = load_index([doc_id])
    chunks = retrieve(question, index, metadata_dict, position_map, top_k=10, rerank_top_n=5)
    return answer_question(question, chunks, chain)


def run_batch(
    question_list: List[str],
    doc_configs: List[Dict[str, str]],
    chain,
    progress_callback: Optional[Callable[[int, int], None]] = None,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Run Q&A for every (question, document) pair.

    Args:
        question_list: List of question strings.
        doc_configs: List of dicts with 'doc_id' and 'filename' keys.
        chain: Chat LLM from build_llm() (Azure or OpenAI — env-driven).
        progress_callback: Optional fn(current, total) for progress tracking.

    Returns:
        (answers_df, trace_df) — answers grid and full audit trace.
    """
    total = len(question_list) * len(doc_configs)
    done = 0

    answers: Dict[str, Dict[str, str]] = {q: {} for q in question_list}
    trace_rows = []

    for question in question_list:
        for doc in doc_configs:
            doc_id = doc["doc_id"]
            filename = doc["filename"]
            try:
                result = _answer_for_doc(question, filename, doc_id, chain)
                answers[question][filename] = result.answer

                top_chunk = result.retrieved_chunks[0] if result.retrieved_chunks else None
                trace_rows.append({
                    "query": question,
                    "filename": filename,
                    "doc_id": doc_id,
                    "answer": result.answer,
                    "confidence_level": result.confidence_level,
                    "top_chunk_page": top_chunk.page_number if top_chunk else "",
                    "top_chunk_section": top_chunk.section_heading if top_chunk else "",
                    "top_reranker_score": top_chunk.reranker_score if top_chunk else "",
                    "prompt_tokens": result.prompt_tokens,
                    "completion_tokens": result.completion_tokens,
                    "latency_seconds": round(result.latency_seconds, 2),
                    "timestamp": result.timestamp,
                    "model_deployment": result.model_deployment,
                    "extraction_methods_used": ",".join(
                        dict.fromkeys(c.extraction_method for c in result.retrieved_chunks)
                    ),
                    "chunk_ids_used": ",".join(
                        c.chunk_id for c in result.retrieved_chunks
                    ),
                })
            except Exception as exc:
                logger.error("Batch error (%s, %s): %s", question[:30], filename, exc)
                answers[question][filename] = f"ERROR: {exc}"
                trace_rows.append({
                    "query": question, "filename": filename, "doc_id": doc_id,
                    "answer": f"ERROR: {exc}", "confidence_level": "Low",
                    "top_chunk_page": "",
                    "top_chunk_section": "",
                    "top_reranker_score": "",
                    "prompt_tokens": 0,
                    "completion_tokens": 0,
                    "latency_seconds": 0.0,
                    "timestamp": "",
                    "model_deployment": "",
                    "extraction_methods_used": "",
                    "chunk_ids_used": "",
                })

            done += 1
            if progress_callback:
                progress_callback(done, total)

    filenames = [d["filename"] for d in doc_configs]
    answers_df = pd.DataFrame(
        {fn: [answers[q].get(fn, "") for q in question_list] for fn in filenames},
        index=question_list,
    )
    trace_df = pd.DataFrame(trace_rows)
    return answers_df, trace_df


def export_to_excel(
    answers_df: pd.DataFrame,
    trace_df: pd.DataFrame,
    output_path: str,
) -> str:
    """Write answers grid and trace log to a formatted Excel workbook.

    Args:
        answers_df: Rows=questions, columns=filenames, cells=answer text.
        trace_df: One row per (question, doc) with all audit metadata.
        output_path: Path to write the .xlsx file.

    Returns:
        output_path (for convenience in the Streamlit download button).
    """
    from openpyxl.styles import PatternFill, Font
    from openpyxl.utils import get_column_letter

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        answers_df.to_excel(writer, sheet_name="Answers", index=True)
        trace_df.to_excel(writer, sheet_name="Trace", index=False)

        ws_ans = writer.sheets["Answers"]
        ws_ans.freeze_panes = "B2"
        header_fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
        alt_fill = PatternFill(fill_type="solid", fgColor="F2F2F2")
        for col_idx, col in enumerate(ws_ans.iter_cols(), start=1):
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws_ans.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 60)
        for row_idx, row in enumerate(ws_ans.iter_rows(), start=1):
            for cell in row:
                if row_idx == 1:
                    cell.fill = header_fill
                    cell.font = Font(bold=True)
                elif row_idx % 2 == 0:
                    cell.fill = alt_fill

    logger.info("Exported batch results to %s", output_path)
    return output_path
